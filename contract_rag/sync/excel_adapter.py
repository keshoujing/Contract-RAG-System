"""Excel ledger I/O boundary (decision 15) — the only module that touches .xlsx.

Append-if-absent / update-system-columns, addressed by the ``contract_id`` column.
A locked workbook (someone has it open) raises :class:`ExcelLocked` rather than
corrupting or blocking — the service degrades that to a ``pending`` state and
retries later. This is the detachable limb: disabling Excel means never calling it.

Columns are addressed by *header name* (row 1) via a field->header mapping, so the
real ledger's headers can differ from our field names without touching logic.
"""
from __future__ import annotations

import pathlib

from openpyxl import Workbook, load_workbook

from contract_rag.sync.models import DEFAULT_EXCEL_COLUMNS, KEY_FIELD, SYNCED_FIELDS


def _normalize_header(value: object) -> str:
    """Collapse all whitespace/newlines so bilingual multi-line headers compare cleanly."""
    return "".join(str(value).split())


class ExcelSyncError(Exception):
    """Base class for ledger I/O failures."""


class ExcelLocked(ExcelSyncError):
    """The workbook is open/locked elsewhere; the write should be retried later."""


class ExcelAdapter:
    """Read/write a single contract row in the ledger workbook.

    Args:
        path: the .xlsx workbook (created with a header row if missing).
        columns: field -> header-name map (defaults to ``DEFAULT_EXCEL_COLUMNS``).
        sheet: worksheet name (defaults to the active sheet).
    """

    def __init__(self, path, columns: dict[str, str] | None = None, sheet: str | None = None):
        self.path = pathlib.Path(path)
        self.columns = columns or dict(DEFAULT_EXCEL_COLUMNS)
        self.sheet = sheet

    # -- internal helpers --------------------------------------------------- #

    def _load(self):
        try:
            wb = load_workbook(self.path)
        except FileNotFoundError:
            wb = self._new_workbook()
        ws = wb[self.sheet] if self.sheet else wb.active
        return wb, ws

    def _new_workbook(self):
        wb = Workbook()
        ws = wb.active
        if self.sheet:
            ws.title = self.sheet
        ws.append([self.columns[f] for f in SYNCED_FIELDS])
        return wb

    def _header_index(self, ws) -> dict[str, int]:
        """Map field name -> 1-based column index by matching header text.

        Real ledger headers are bilingual and multi-line (e.g. ``"Contract No.\n
        (合同编号)"``), so matching is whitespace-insensitive and substring-based:
        a mapped value matches a header if, after stripping all whitespace, it is
        contained in the header. Pick distinctive labels in ``self.columns`` (e.g.
        the Chinese token) — but where two headers share one (``存档编号`` appears
        in both "File No." and "File Name"), use an English token to disambiguate.
        """
        headers = [
            (_normalize_header(cell.value), cell.column)
            for cell in ws[1]
            if cell.value is not None
        ]
        idx: dict[str, int] = {}
        for f, label in self.columns.items():
            needle = _normalize_header(label)
            if not needle:
                continue
            for header, col in headers:
                if needle in header:
                    idx[f] = col
                    break
        return idx

    def _save(self, wb) -> None:
        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(self.path)
        except PermissionError as e:  # workbook open in Excel -> retry later
            raise ExcelLocked(f"ledger locked: {self.path}") from e

    # -- public API --------------------------------------------------------- #

    def find_row(self, contract_id: str) -> dict | None:
        """Return the ledger row as a field->value dict, or None if absent."""
        if not self.path.exists():
            return None
        _, ws = self._load()
        idx = self._header_index(ws)
        key_col = idx.get(KEY_FIELD)
        if key_col is None:
            return None
        for row in ws.iter_rows(min_row=2):
            if str(row[key_col - 1].value) == str(contract_id):
                return {f: row[c - 1].value for f, c in idx.items()}
        return None

    def upsert_row(self, contract_id: str, values: dict[str, object]) -> None:
        """Append a new row, or update only the given fields of an existing row.

        ``values`` is field->value (the key field is forced to ``contract_id``).
        Raises :class:`ExcelLocked` if the workbook cannot be written.
        """
        wb, ws = self._load()
        idx = self._header_index(ws)
        key_col = idx.get(KEY_FIELD)
        if key_col is None:
            raise ExcelSyncError(f"ledger has no '{self.columns[KEY_FIELD]}' column")

        target = None
        for row in ws.iter_rows(min_row=2):
            if str(row[key_col - 1].value) == str(contract_id):
                target = row
                break

        payload = {**values, KEY_FIELD: contract_id}
        if target is None:
            # Append a FULL-WIDTH row so every existing column is preserved
            # (decision 15: the ledger format never changes — unmapped columns
            # stay blank, none are dropped or reordered).
            width = ws.max_column
            row_vals: list[object | None] = [None] * width
            for f, value in payload.items():
                col = idx.get(f)
                if col is not None and col <= width:
                    row_vals[col - 1] = value
            ws.append(row_vals)
        else:
            for f, value in payload.items():
                col = idx.get(f)
                if col is not None:
                    target[col - 1].value = value
        self._save(wb)
