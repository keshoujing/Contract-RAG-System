"""Excel adapter: preserve the ledger's full column layout + tolerant header match.

Guards decision 15's hard constraint — the ledger format never changes: appends
are full-width, unmapped/human columns are never dropped or reordered, and
bilingual multi-line headers (incl. the ``存档编号`` collision) resolve correctly.
"""
from __future__ import annotations

from openpyxl import Workbook, load_workbook

from contract_rag.sync.excel_adapter import ExcelAdapter

# A realistic slice of the real ledger: bilingual multi-line headers, a spacer
# column, and an unmapped human column ("Remarks").
_HEADERS = [
    "File No.\n(存档编号)",
    "",                       # spacer column (blank header)
    "Supplier\n(供应商)",
    "Contract No.\n(合同编号)",
    "Contract Amount\n(合同金额)",
    "Remarks\n（备注）",       # unmapped: only humans write here
]
_COLS = {"contract_id": "合同编号", "counterparty": "供应商", "amount": "合同金额"}


def _ledger(tmp_path):
    p = tmp_path / "ledger.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(_HEADERS)
    ws.append(["F-1", None, "Old Supplier", "C-1", 100, "人工备注"])  # pre-existing human row
    wb.save(p)
    return p


def test_append_is_full_width_and_keeps_all_columns(tmp_path):
    p = _ledger(tmp_path)
    ExcelAdapter(p, columns=_COLS).upsert_row("C-2", {"counterparty": "Jushi Egypt", "amount": 39041.6})

    ws = load_workbook(p).active
    assert ws.max_column == 6  # no columns dropped
    new = [c.value for c in ws[3]]
    assert new[2] == "Jushi Egypt"  # Supplier
    assert new[3] == "C-2"          # Contract No.
    assert new[4] == 39041.6        # Contract Amount
    assert new[0] is None and new[1] is None and new[5] is None  # unmapped cols stay blank


def test_update_touches_only_mapped_cells(tmp_path):
    p = _ledger(tmp_path)
    ExcelAdapter(p, columns=_COLS).upsert_row("C-1", {"amount": 999})

    ws = load_workbook(p).active
    row = [c.value for c in ws[2]]
    assert row[4] == 999          # amount updated
    assert row[2] == "Old Supplier"  # untouched
    assert row[5] == "人工备注"      # human-only column preserved


def test_header_collision_disambiguated_by_english_token(tmp_path):
    p = tmp_path / "l.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["File No.\n(存档编号)", "File Name\n(存档编号)", "Contract No.\n(合同编号)"])
    wb.save(p)

    # "存档编号" alone is ambiguous; English tokens disambiguate the two columns.
    adapter = ExcelAdapter(p, columns={"file_no": "File No.", "file_name": "File Name"})
    idx = adapter._header_index(load_workbook(p).active)
    assert idx["file_no"] == 1
    assert idx["file_name"] == 2
